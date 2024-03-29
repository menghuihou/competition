
containers = [11, 5, 4, 7, 16, 6, 5, 7,
     13, 6, 5, 7, 12, 5, 4, 6,
     9, 5, 5, 11, 29, 21, 17, 20,
     27, 13, 9, 10, 16, 6, 5, 7,
     11, 5, 5, 6, 12, 7, 7, 10,
     15, 10, 9, 11, 15, 10, 10, 16,
     26, 21, 23, 36, 50, 45, 45, 49,
     57, 43, 40, 44, 52, 43, 42, 45,
     52, 41, 39, 41, 48, 35, 34, 35,
     42, 34, 36, 43, 55, 48, 54, 65,
     80, 70, 74, 85, 101, 89, 88, 90,
     100, 87, 88, 89, 104, 89, 89, 90,
     106, 96, 94, 99, 109, 99, 96, 102]

arms = []
for i in range(len(containers)):
    arms.append(4*containers[i])

addContainers = []

import openpyxl
wb = openpyxl.load_workbook('E:\\hmh\\D\\D22\\Python\\PyCharm\\My Project\\python\\数学建模\\问题三：1-104周的结果数据.xlsx')
sheet = wb.active

row_num = sheet.max_row
for row in range(2,row_num+1):
    cell = sheet.cell(row,2)
    addContainers.append(cell.value)


containersMaintain = []
costX = []
for i in range(0,104):
     containersSum = 0
     containersCost = 0
     for j in range(len(addContainers)):
          containersSum = containersSum + addContainers[j]
          if(containersSum <= 5):
               containersCost = containersSum * 200
          elif(5 < containersSum <= 10):
               containersCost = 5 * 200 + (containersSum - 5) * 180
          elif(containersSum > 10):
               containersCost = 5 * 200 + 5 * 180 +(containersSum - 10) * 160
          #containersCost += containersCost
          costX.append(containersCost)

#print(costX)


for i in range(0,104):
     maintance = []
     containersSum = 0
     containersCost = 0
     for j in range(i,104):
          containersSum = containersSum + addContainers[j]
          maintance.append(maintance + addContainers[j] * j * 10)
          if (containersSum <= 5):
               containersCost = containersSum * 200
          elif (5 < containersSum <= 10):
               containersCost = 5 * 200 + (containersSum - 5) * 180
          elif (containersSum > 10):
               containersCost = 5 * 200 + 5 * 180 + (containersSum - 10) * 160
          costX.append(containersCost)
          containersMaintain.append(maintance)

print(costX)
print(maintance)

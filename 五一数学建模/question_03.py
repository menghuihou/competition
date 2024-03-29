import math
from openpyxl import  Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

container = [11, 5, 4, 7, 16, 6, 5, 7,
             13, 6, 5, 7, 12, 5, 4, 6,
             9, 5, 5, 11, 29, 21, 17, 20,
             27, 13, 9, 10, 16, 6, 5, 7,
             11, 5, 5, 6, 12, 7, 7, 10,
             15, 10, 9, 11, 15, 10, 10, 16,
             26, 21, 23, 36, 50, 45, 45, 49,
             57, 43, 40, 44, 52, 43, 42, 45,
             52, 41, 39, 41, 48, 35, 34, 35,
             42, 34, 36, 43, 55, 48, 54, 65,
             80, 70, 74, 85, 101, 89, 88, 90,
             100, 87, 88, 89, 104, 89, 89, 90,
             106, 96, 94, 99, 109, 99, 96, 102]  # 容器艇

arm = []  # 机械臂
for i in range(len(container)):
    arm.append(container[i] * 4)

initial_arm = 50  # 初始机械臂数量
initial_container = 13  # 初始容器艇数量
weeks_arm = []  # 存储需要购买机械臂的周数
weeks_container = []  # 存储需要购买容器艇的周数
maintenance_costs_arm, maintenance_costs_container = 0, 0
num1 = 0
num2 = 0
add_arm = []  # 新增的机械臂数量
add_container = []  # 新增的容器艇数量
weekly_damage_arm = []  # 每周损毁的机械臂数量
weekly_damage_container = []  # 每周损毁的容器艇数量

weekly_maintenance_arm = []  # 每周需要保养的机械臂数量
weekly_maintenance_container = []  # 每周需要保养的容器艇数量


v = []  # 截止第n周的总成本
total_cost = 0  # 总成本
a = []
for i in range(len(container)):
    weekly_damage_container.append(int(container[i] * 0.1 + 0.5))
    weekly_damage_arm.append(4 * weekly_damage_container[i])
    if (i + 1 < len(arm)):
        s = arm[i] + arm[i + 1]
        weekly_maintenance_arm.append(initial_arm - arm[i])  # 第i周所需保养的机械手数量
        if (s > initial_arm):  # 机械臂
            num1 = (s - initial_arm)
            add_arm.append(num1)
            a = math.ceil(num1 / 20)  # 每周参与训练的熟练工数量
            v.append(num1 + a)  # 参与训练的操作手数量
            initial_arm = s
            weeks_arm.append(i)
        else:
            a = 0
            add_arm.append(0)
            v.append(0)
        weekly_maintenance_arm[i] = weekly_maintenance_arm[i] - a  # 第i周所需保养的机械手数量
    initial_arm = initial_arm - weekly_damage_arm[i]
    print(weekly_maintenance_arm)
    if (container[i] > initial_container):  # 容器艇
        num2 = container[i] - initial_container
        add_container.append(num2)
        initial_container = container[i]
        weeks_container.append(i - 1)
    else:
        add_container.append(0)
    weekly_maintenance_container.append(initial_container - container[i])  # 第i周所需保养的容器艇数量
   # initial_container = initial_container - weekly_damage_container[i]
    if (i + 1 >= len(container)):
        break
    else:
        maintenance_costs_arm = maintenance_costs_arm + weekly_maintenance_arm[i] * 5  # 机械臂的保养费用

    maintenance_costs_container = maintenance_costs_container + (initial_container - container[i]) * 10  # 容器艇的保养费用
    initial_container = initial_container - weekly_damage_container[i]
maintenance_costs_arm = maintenance_costs_arm + (initial_arm+weekly_damage_arm[103]-arm[103])*5  # 机械臂的保养费用
v.append(0)
add_arm.append(0)
# add_container.append(0) #另add_arm和add_container长度为104
add_container.remove(add_container[0])
add_container.append(0)
weekly_maintenance_arm.append(0)

for i in range(len(add_arm)):
    print("第", i + 1, "周购买的机械臂数量为：", add_arm[i])
for i in range(len(add_container)):
    print("第", i+1, "周购买的容器艇数量为：", add_container[i])


total_add_arm = sum(add_arm)  # 新增的机械臂总和
expenses_arm = sum(add_arm) * 100  # 机械臂的购买费用
expenses_container = sum(add_container) * 200  # 容器艇的购买费用
training_expenses_arm = sum(v) * 10  # 训练机械臂的费用
total_cost = expenses_arm + expenses_container + training_expenses_arm + maintenance_costs_arm + maintenance_costs_container

print("总成本：", total_cost, "元")

#周数（104周）
weeks = []
for i in range(0, 104):
    weeks1 = "第%d周"% (i+1)
    weeks.append(weeks1)

#参与训练的操作手数量
training_arm = []
for i in range(len(add_arm)):
    training_arm_num = math.ceil(add_arm[i]/10)
    training_arm.append(training_arm_num)

# 绘制表格
wb = Workbook()
sheet = wb.active
sheet.title='问题三：1-104周的结果数据'
sheet.append(['周数', '购买的容器艇数量', '购买的操作手数量','保养的操作手数量','保养的容器艇数量','参与训练的操作手数量','总成本（单位：元）'])
for i in range(0,104):
    sheet.append([weeks[i],add_container[i],add_arm[i],weekly_maintenance_arm[i],weekly_maintenance_container[i],v[i]])
wb.save('问题三：1-104周的结果数据.xlsx')
wb = load_workbook('问题三：1-104周的结果数据.xlsx')
sheet = wb.active

#表格列宽自适应
# 第一步：计算每列最大宽度，并存储在列表lks中。
lks = []  # 英文变量太费劲，用汉语首字拼音代替
for i in range(1, sheet.max_column + 1):  # 每列循环
    lk = 1  # 定义初始列宽，并在每个行循环完成后重置
    for j in range(1, sheet.max_row + 1):  # 每行循环
        sz = sheet.cell(row=j, column=i).value  # 每个单元格内容
        if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
            lk1 = len(sz.encode('gbk'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
        else:
            lk1 = len(str(sz))
        if lk < lk1:
            lk = lk1  # 借助每行循环将最大值存入lk中
        # print(lk)
    lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）
# 第二步：设置列宽
for i in range(1, sheet.max_column + 1):
    k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
    sheet.column_dimensions[k].width = lks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整
wb.close()
wb.save('问题三：1-104周的结果数据.xlsx')

print("购买的操作手数量：",sum(add_arm))
print("购买的容器艇数量：",sum(add_container))


print(len(add_container))
print(initial_arm+weekly_damage_arm[103]-arm[103])